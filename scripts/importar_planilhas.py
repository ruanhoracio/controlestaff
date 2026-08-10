#!/usr/bin/env python3
"""
Lê as três planilhas da Maxipas e gera supabase/importacao.sql.

Uso:
    python3 scripts/importar_planilhas.py "/caminho/para/pasta-com-as-planilhas"

O SQL gerado descobre o user_id pelo e-mail (constante EMAIL abaixo), então
não é preciso saber o UUID. Rode-o no SQL Editor do Supabase depois da
migração 01.
"""
from __future__ import annotations

import datetime
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

EMAIL = "Karolayne.Silveiraka@gmail.com"

ARQ_PPP = "Controle de PPP.xlsx"
ARQ_CAT = "Abertura de CAT - Célula .xlsx"
ARQ_ROD = "RODÍZIO - DESIGNAR CÉLULA.xlsx"

MESES_PT = {
    "JANEIRO": 1, "JAN": 1, "FEVEREIRO": 2, "FEV": 2, "MARÇO": 3, "MARCO": 3, "MAR": 3,
    "ABRIL": 4, "ABR": 4, "MAIO": 5, "MAI": 5, "JUNHO": 6, "JUN": 6, "JULHO": 7, "JUL": 7,
    "AGOSTO": 8, "AGO": 8, "SETEMBRO": 9, "SET": 9, "OUTUBRO": 10, "OUT": 10,
    "NOVEMBRO": 11, "NOV": 11, "DEZEMBRO": 12, "DEZ": 12,
}

CONCLUSOES = {
    "ENTREGUE": "ENTREGUE",
    "PENDENTE": "PENDENTE",
    "AUXILIO": "AUXILIO",
    "NAO SE APLICA": "NAO_SE_APLICA",
    "DESCONSIDERADO": "DESCONSIDERADO",
}

# Blocos de colunas da aba atual do rodízio -> id do rodízio no app
BLOCOS_RODIZIO_ATUAL = [
    (1, "Empreiteiras Regionais"),
    (7, "Pequena Empresa"),
    (13, "Média Empresa"),
    (19, "Grande Empresa"),
    (25, "Rede Corporativa"),
    (31, "Exames, Pontual, PGSM e Licitações"),
    (37, "Ergonomistas"),
]

SIGLA_PORTE = {"PE": "Pequena Empresa", "ME": "Média Empresa", "GE": "Grande Empresa", "RC": "Rede Corporativa"}

avisos: list[str] = []


# ---------------------------------------------------------------- utilidades


def sem_acento(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def norm(v) -> str:
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def data_iso(v) -> str | None:
    """Converte célula em 'YYYY-MM-DD'. Devolve None quando não é data."""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    t = norm(v)
    # tolera espaços em volta das barras ('21/08 /2024' aparece na planilha)
    m = re.match(r"^(\d{1,2})\s*/\s*(\d{1,2})\s*/\s*(\d{2,4})$", t)
    if m:
        d, mo, a = (int(x) for x in m.groups())
        if a < 100:
            a += 2000
        try:
            return datetime.date(a, mo, d).isoformat()
        except ValueError:
            return None
    return None


def todas_as_datas(v) -> list[str]:
    """Extrai todas as datas de uma célula ('09/07/2026 - 23/07/2026' -> duas)."""
    uma = data_iso(v)
    if uma:
        return [uma]
    t = norm(v)
    achadas = []
    for m in re.finditer(r"(\d{1,2})\s*/\s*(\d{1,2})\s*/\s*(\d{2,4})", t):
        d, mo, a = (int(x) for x in m.groups())
        if a < 100:
            a += 2000
        try:
            achadas.append(datetime.date(a, mo, d).isoformat())
        except ValueError:
            pass
    return achadas


def sql(v) -> str:
    if v is None or v == "":
        return "null" if v is None else "''"
    return "'" + str(v).replace("'", "''") + "'"


def dias_uteis(inicio: str, n: int) -> str:
    d = datetime.date.fromisoformat(inicio)
    somados = 0
    while somados < n:
        d += datetime.timedelta(days=1)
        if d.weekday() < 5:
            somados += 1
    return d.isoformat()


# ------------------------------------------------------------------- módulo 1


def ler_ppp(pasta: Path) -> list[dict]:
    ws = openpyxl.load_workbook(pasta / ARQ_PPP, data_only=True)["Controle de PPP"]
    registros, ignorados = [], Counter()

    for r in range(2, ws.max_row + 1):
        empresa, funcionario = norm(ws.cell(r, 1).value), norm(ws.cell(r, 2).value)
        if not empresa and not funcionario:
            continue
        # linhas de cabeçalho repetidas no meio da planilha
        if empresa.upper() == "EMPRESA" or funcionario.upper() == "NOME":
            ignorados["cabeçalho repetido"] += 1
            continue

        bruta = sem_acento(norm(ws.cell(r, 12).value)).upper()
        conclusao = CONCLUSOES.get(bruta)
        if conclusao is None:
            if bruta:
                ignorados[f"conclusão desconhecida: {bruta}"] += 1
            conclusao = "PENDENTE"

        entrega = data_iso(ws.cell(r, 9).value)
        solicitado = data_iso(ws.cell(r, 7).value)
        prazo = data_iso(ws.cell(r, 8).value)
        observacao = norm(ws.cell(r, 13).value)

        # Algumas linhas têm a fórmula quebrada ('#VALUE!' ou o número 7 cru).
        # Quando há data de entrega, ela salva o registro; senão não dá pra situar no tempo.
        if not solicitado and not prazo:
            if entrega:
                solicitado = prazo = entrega
                nota = "datas de solicitação e prazo ilegíveis na planilha de origem"
                observacao = f"{observacao} · {nota}".strip(" ·")
                avisos.append(f"PPP: '{funcionario}' ({empresa}) importado usando a data de entrega — {nota}")
            else:
                avisos.append(
                    f"PPP: '{funcionario}' ({empresa or 'sem empresa'}) NÃO importado — "
                    "linha sem empresa e sem nenhuma data. Precisa ser recadastrada no app."
                )
                ignorados["sem nenhuma data aproveitável"] += 1
                continue
        elif not solicitado:
            solicitado = prazo
        elif not prazo:
            prazo = dias_uteis(solicitado, 7)

        mes = (entrega if conclusao == "ENTREGUE" and entrega else prazo)[:7]

        registros.append(
            dict(
                empresa=empresa,
                funcionario=funcionario,
                celula=norm(ws.cell(r, 3).value),
                tipo=norm(ws.cell(r, 4).value),
                admissao=data_iso(ws.cell(r, 5).value),
                demissao=data_iso(ws.cell(r, 6).value),  # 'ATIVO' vira null
                data_solicitada=solicitado,
                prazo_entrega=prazo,
                data_entrega=entrega,
                responsavel=norm(ws.cell(r, 10).value),
                mes=mes,
                conclusao=conclusao,
                observacao=observacao,
            )
        )

    for motivo, n in ignorados.items():
        avisos.append(f"PPP: {n} linha(s) ignorada(s) — {motivo}")
    return registros


# ------------------------------------------------------------------- módulo 2


def ler_rodizio(pasta: Path) -> list[dict]:
    """
    A vez avança pela ORDEM DAS LINHAS da planilha, não pela data: as datas
    não são monotônicas (há 12/05 seguido de 07/05 em Empreiteiras). Por isso
    'ordem' é a posição na fila, e é ela que vira o created_at no banco.
    """
    wb = openpyxl.load_workbook(pasta / ARQ_ROD, data_only=True)
    designacoes, ignorados = [], Counter()

    # --- aba antiga primeiro: é o histórico mais velho ---
    wsa = wb["RODÍZIO CÉLULA"]
    for col in (1, 7):
        for r in range(2, wsa.max_row + 1):
            posicao = norm(wsa.cell(r, col).value)
            if not posicao or not posicao.upper().startswith("CÉLULA"):
                continue
            data = data_iso(wsa.cell(r, col + 3).value)
            if not data:
                continue
            sigla = norm(wsa.cell(r, col + 2).value).upper()
            porte = SIGLA_PORTE.get(sigla)
            if not porte:
                ignorados["aba antiga, porte não identificado"] += 1
                continue
            designacoes.append(
                dict(
                    data=data,
                    porte=porte,
                    empresa=None,
                    celula=normalizar_celula(posicao),
                    responsavel=norm(wsa.cell(r, col + 1).value).title(),
                    esocial="",
                    pulada=False,
                )
            )

    # --- aba atual: sete rodízios lado a lado ---
    ws = wb["RODÍZIO CÉLULA - 31102025"]
    for col, rodizio in BLOCOS_RODIZIO_ATUAL:
        for r in range(3, ws.max_row + 1):
            posicao = norm(ws.cell(r, col).value)
            if not posicao:
                continue
            celula_val = ws.cell(r, col + 3).value
            marca = norm(ws.cell(r, col + 4).value)
            data = data_iso(celula_val)
            pulada = data is None and norm(celula_val) in {"--", "-"}
            if data is None and not pulada:
                continue  # ainda não usada: é a fila futura, não histórico
            pessoa = norm(ws.cell(r, col + 1).value).title()
            ergonomista = rodizio == "Ergonomistas"
            designacoes.append(
                dict(
                    data=data or ultima_data(designacoes, rodizio),
                    porte=rodizio,
                    empresa=None,
                    # No rodízio de ergonomistas a fila são as pessoas, não as células
                    celula=pessoa if ergonomista else normalizar_celula(posicao),
                    responsavel=pessoa,
                    esocial="" if ergonomista else norm(ws.cell(r, col + 2).value),
                    pulada=pulada,
                )
            )
            _ = marca

    for motivo, n in ignorados.items():
        avisos.append(f"Rodízio: {n} linha(s) ignorada(s) — {motivo}")

    # A posição na lista JÁ é a ordem da fila — não reordenar por data.
    return designacoes


def ultima_data(designacoes: list[dict], rodizio: str) -> str:
    """Data a usar numa vez pulada: a da designação anterior do mesmo rodízio."""
    for d in reversed(designacoes):
        if d["porte"] == rodizio:
            return d["data"]
    return "2025-11-01"


def normalizar_celula(v: str) -> str:
    """'CÉLULA I' -> 'Célula I'. Nomes de ergonomistas passam intactos."""
    if v.upper().startswith("CÉLULA") or v.upper().startswith("CELULA"):
        return "Célula " + v.split()[-1].upper()
    return v.title()


# ------------------------------------------------------------------- módulo 3


def ler_cat(pasta: Path) -> tuple[list[dict], dict, dict]:
    wb = openpyxl.load_workbook(pasta / ARQ_CAT, data_only=True)
    cats, ignorados = [], Counter()
    equipe_por_mes: dict[str, dict[str, list[str]]] = {}
    gestores: dict[str, str] = {}

    for aba in wb.sheetnames:
        m = re.match(r"^([A-Za-zÇç]+)(\d{4})$", aba.strip())
        if not m:
            continue
        mes_num = MESES_PT.get(sem_acento(m.group(1)).upper().replace("C", "C"))
        if mes_num is None:
            mes_num = MESES_PT.get(m.group(1).upper())
        if mes_num is None:
            avisos.append(f"CAT: aba '{aba}' ignorada — mês não reconhecido")
            continue
        ano = int(m.group(2))
        ws = wb[aba]
        equipe_mes: dict[str, list[str]] = defaultdict(list)

        for r in range(1, ws.max_row + 1):
            tecnico = norm(ws.cell(r, 1).value)
            celula_gestor = norm(ws.cell(r, 2).value)
            if not tecnico or tecnico.upper().startswith("TÉCNICO"):
                continue
            if not celula_gestor or "célula" not in celula_gestor.lower():
                continue

            partes = [p.strip() for p in celula_gestor.split("-", 1)]
            celula = normalizar_celula(partes[0])
            if len(partes) > 1 and partes[1]:
                gestores[celula] = partes[1]
            if tecnico not in equipe_mes[celula]:
                equipe_mes[celula].append(tecnico)

            for data in todas_as_datas(ws.cell(r, 3).value):
                # Data fora do mês da aba: vale a data escrita, que é quando o CAT
                # de fato aconteceu — a aba errada é só onde foi anotado.
                if data[:7] != f"{ano:04d}-{mes_num:02d}":
                    avisos.append(
                        f"CAT: '{tecnico}' anotado na aba {aba} com data {data} — "
                        "importado no mês da data, não no da aba"
                    )
                cats.append(dict(data=data, celula=celula, tecnico=tecnico, direta=False))

            if norm(ws.cell(r, 3).value) and not todas_as_datas(ws.cell(r, 3).value):
                ignorados[f"data ilegível ({aba})"] += 1

        if equipe_mes:
            equipe_por_mes[f"{ano:04d}-{mes_num:02d}"] = dict(equipe_mes)

    for motivo, n in ignorados.items():
        avisos.append(f"CAT: {n} registro(s) ignorado(s) — {motivo}")

    ultimo = max(equipe_por_mes) if equipe_por_mes else None
    equipe_atual = equipe_por_mes.get(ultimo, {}) if ultimo else {}
    cats.sort(key=lambda c: c["data"])
    return cats, equipe_atual, gestores


# ------------------------------------------------------------------- geração


def gerar_sql(ppps, designacoes, cats, equipe_cat, gestores, esocial, ergonomistas) -> str:
    import json

    celulas = []
    for nome in ["Célula I", "Célula II", "Célula III"]:
        celulas.append(
            {
                "nome": nome,
                "responsavel": gestores.get(nome, ""),
                "gestor": gestores.get(nome, ""),
                "tecnicos": sorted(equipe_cat.get(nome, []), key=lambda s: s.lower()),
            }
        )
    equipe = {"celulas": celulas, "esocial": esocial, "ergonomistas": ergonomistas}

    L = []
    A = L.append
    A("-- ============================================================")
    A("-- Importação das planilhas da Maxipas para o Controle Staff")
    A("-- GERADO AUTOMATICAMENTE por scripts/importar_planilhas.py")
    A("--")
    A(f"--   {len(ppps):5d} registros de PPP")
    A(f"--   {len(designacoes):5d} designações de célula")
    A(f"--   {len(cats):5d} designações de CAT")
    A("--")
    A("-- Rode a migracao-01-planilhas.sql ANTES deste arquivo.")
    A("-- É idempotente: apaga os dados da conta e reimporta do zero.")
    A("-- ============================================================")
    A("")
    A("do $$")
    A("declare")
    A("  uid uuid;")
    A("begin")
    A(f"  select id into uid from auth.users where lower(email) = lower({sql(EMAIL)});")
    A("  if uid is null then")
    A(f"    raise exception 'Usuário % não encontrado em auth.users', {sql(EMAIL)};")
    A("  end if;")
    A("")
    A("  delete from public.ppp_records        where user_id = uid;")
    A("  delete from public.designacoes_celula where user_id = uid;")
    A("  delete from public.designacoes_cat    where user_id = uid;")
    A("")
    A("  -- ---------------- Equipe ----------------")
    A("  update public.app_config set equipe = " + sql(json.dumps(equipe, ensure_ascii=False)) + "::jsonb")
    A("   where user_id = uid;")
    A("  insert into public.app_config (user_id, equipe)")
    A("  select uid, " + sql(json.dumps(equipe, ensure_ascii=False)) + "::jsonb")
    A("   where not exists (select 1 from public.app_config where user_id = uid);")
    A("")

    if ppps:
        A("  -- ---------------- Módulo 1: PPP ----------------")
        A("  insert into public.ppp_records (user_id, empresa, funcionario, celula, tipo, admissao,")
        A("    demissao, data_solicitada, prazo_entrega, data_entrega, responsavel, mes, conclusao, observacao)")
        A("  values")
        linhas = [
            "    (uid, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {})".format(
                sql(p["empresa"]), sql(p["funcionario"]), sql(p["celula"]), sql(p["tipo"]),
                sql(p["admissao"]), sql(p["demissao"]), sql(p["data_solicitada"]),
                sql(p["prazo_entrega"]), sql(p["data_entrega"]), sql(p["responsavel"]),
                sql(p["mes"]), sql(p["conclusao"]), sql(p["observacao"]),
            )
            for p in ppps
        ]
        A(",\n".join(linhas) + ";")
        A("")

    if designacoes:
        A("  -- ---------------- Módulo 2: rodízio de célula ----------------")
        A("  -- created_at é uma sequência sintética que preserva a ORDEM DA FILA da planilha.")
        A("  -- É ela que o app usa para saber de quem é a vez; a coluna 'data' guarda a data real.")
        A("  insert into public.designacoes_celula (user_id, data, porte, empresa, celula,")
        A("    responsavel, esocial, pulada, created_at)")
        A("  values")
        linhas = [
            "    (uid, {}, {}, {}, {}, {}, {}, {}, {}::timestamptz)".format(
                sql(d["data"]), sql(d["porte"]), sql(d["empresa"]), sql(d["celula"]),
                sql(d["responsavel"]), sql(d["esocial"]), "true" if d["pulada"] else "false",
                sql((datetime.datetime(2020, 1, 1) + datetime.timedelta(minutes=i)).isoformat() + "Z"),
            )
            for i, d in enumerate(designacoes)
        ]
        A(",\n".join(linhas) + ";")
        A("")

    if cats:
        A("  -- ---------------- Módulo 3: rodízio de CAT ----------------")
        A("  insert into public.designacoes_cat (user_id, data, celula, tecnico, direta, created_at)")
        A("  values")
        linhas = [
            "    (uid, {}, {}, {}, {}, {}::timestamptz)".format(
                sql(c["data"]), sql(c["celula"]), sql(c["tecnico"]),
                "true" if c["direta"] else "false",
                sql(f"{c['data']}T12:{i // 60:02d}:{i % 60:02d}Z"),
            )
            for i, c in enumerate(cats)
        ]
        A(",\n".join(linhas) + ";")
        A("")

    A("  raise notice 'Importação concluída para %', uid;")
    A("end $$;")
    A("")
    return "\n".join(L)


def main():
    pasta = Path(sys.argv[1] if len(sys.argv) > 1 else ".").expanduser()

    ppps = ler_ppp(pasta)
    designacoes = ler_rodizio(pasta)
    cats, equipe_cat, gestores = ler_cat(pasta)

    # Filas lidas da aba atual do rodízio
    wb = openpyxl.load_workbook(pasta / ARQ_ROD, data_only=True)
    ws = wb["RODÍZIO CÉLULA - 31102025"]
    esocial: list[str] = []
    for r in range(3, ws.max_row + 1):
        nome = norm(ws.cell(r, 45).value)  # bloco eSocial, coluna Responsável
        if nome and nome not in esocial:
            esocial.append(nome)
    ergonomistas: list[dict] = []
    for r in range(3, ws.max_row + 1):
        nome, cel = norm(ws.cell(r, 38).value), norm(ws.cell(r, 37).value)
        if nome and not any(e["nome"] == nome.title() for e in ergonomistas):
            ergonomistas.append({"nome": nome.title(), "celula": normalizar_celula(cel)})

    destino = Path(__file__).resolve().parent.parent / "supabase" / "importacao.sql"
    destino.write_text(gerar_sql(ppps, designacoes, cats, equipe_cat, gestores, esocial, ergonomistas), "utf-8")

    print(f"PPP .............. {len(ppps)} registros")
    print(f"Designações célula {len(designacoes)}")
    print(f"Designações CAT .. {len(cats)}")
    print(f"eSocial .......... {esocial}")
    print(f"Ergonomistas ..... {[e['nome'] for e in ergonomistas]}")
    print(f"Gestores ......... {gestores}")
    for celula, tecnicos in sorted(equipe_cat.items()):
        print(f"  {celula}: {len(tecnicos)} técnicos — {', '.join(sorted(tecnicos))}")
    if avisos:
        print("\nAVISOS:")
        for a in avisos:
            print("  -", a)
    print(f"\nGerado: {destino}  ({destino.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()
